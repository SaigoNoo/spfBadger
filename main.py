from textual.app import App, ComposeResult
from textual.widgets import Header, Footer, Static
from textual.containers import VerticalScroll
from textual.reactive import reactive
from datetime import datetime
import asyncio
from pythonbeid.card_reader import CardReader
from openpyxl import load_workbook, Workbook
from rich.console import Console
from rich.table import Table

console = Console()

# ---------------- EncodePassage ---------------- #
class EncodePassage:
    def add_passage(self, card_id: str):
        file = "passages.xlsx"
        try:
            wb = load_workbook(file)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Heure", "ID Carte", "Portail"])
        ws.append([self.get_time(), card_id, "Gate1"])
        wb.save(file)

    @staticmethod
    def get_time():
        return datetime.now().strftime("%H:%M:%S")

    @staticmethod
    def show_passages():
        """Affiche l'historique complet avec Rich"""
        try:
            wb = load_workbook("passages.xlsx")
            ws = wb.active
            table = Table(title="Historique des passages", show_lines=True)
            for col in ws[1]:
                table.add_column(col.value, justify="center")
            for row in ws.iter_rows(min_row=2, values_only=True):
                table.add_row(*[str(x) if x is not None else "" for x in row])
            console.clear()
            console.print(table)
        except FileNotFoundError:
            console.print("[red]⚠️ Aucun fichier de passages trouvé.[/red]")

    @staticmethod
    def get_history():
        """Retourne toutes les lignes pour le dashboard"""
        file = "passages.xlsx"
        try:
            wb = load_workbook(file)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            return rows
        except FileNotFoundError:
            return [["⚠ Aucun fichier de passages trouvé."]]

# ---------------- Card ---------------- #
class Card:
    def __init__(self):
        self.raw = None
        self.encode = EncodePassage()

    def read_card(self) -> bool:
        try:
            reader = CardReader()
            self.raw = reader.read_informations(photo=False)
            return True
        except RuntimeError:
            return False

    async def wait_for_card(self):
        first = True
        while not self.read_card():
            if first:
                yield "status", "Veuillez insérer une carte..."
                first = False
            await asyncio.sleep(1)
        yield "status", "Lecture en cours..."
        await asyncio.sleep(1)
        yield "card_id", str(self.raw["num_carte"])
        yield "status", "Veuillez retirer votre carte..."


    async def wait_for_remove_card(self):
        while True:
            try:
                if not self.read_card():
                    break
            except:
                break
            await asyncio.sleep(0.5)
        yield "status", "Carte retirée"
        await asyncio.sleep(1)

# ---------------- Dashboard ---------------- #
class CardDashboard(App):
    card_id = reactive("Aucune carte détectée")
    status = reactive("En attente...")
    history_lines = reactive([])

    def __init__(self):
        super().__init__()
        self.card_reader = Card()

    async def monitor_cards(self):
        while True:
            # Lecture de la carte
            async for event, value in self.card_reader.wait_for_card():
                if event == "status":
                    self.status = value
                elif event == "card_id":
                    self.card_id = value
                    self.card_reader.encode.add_passage(value)
                    self.history_lines = self.card_reader.encode.get_history()

            # Retrait de la carte
            async for event, value in self.card_reader.wait_for_remove_card():
                if event == "status":
                    self.status = value
            await asyncio.sleep(0.1)

    def compose(self) -> ComposeResult:
        yield Header()
        yield Footer()
        # Panels principaux
        yield Static(f"Carte: {self.card_id}\nStatus: {self.status}", id="card_panel")
        yield VerticalScroll(Static("\n".join([" | ".join(line) for line in self.history_lines]), id="history_panel"))

    def on_mount(self):
        self.card_panel = self.query_one("#card_panel", Static)
        self.history_panel = self.query_one("#history_panel", Static)
        asyncio.create_task(self.monitor_cards())

    # Updates réactives
    def watch_card_id(self, value: str):
        if hasattr(self, "card_panel"):
            self.card_panel.update(f"Carte: {self.card_id}\nStatus: {self.status}")

    def watch_status(self, value: str):
        if hasattr(self, "card_panel"):
            self.card_panel.update(f"Carte: {self.card_id}\nStatus: {self.status}")

    def watch_history_lines(self, value):
        if hasattr(self, "history_panel"):
            self.history_panel.update("\n".join([" | ".join(line) for line in value]))

if __name__ == "__main__":
    CardDashboard().run()
